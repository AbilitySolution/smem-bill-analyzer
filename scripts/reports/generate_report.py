#!/usr/bin/env python3
"""
Générateur de rapports Excel Ability (démo SMEM) — version simplifiée.

Usage : python3 generate_report.py '<params-json>' <sortie.xlsx>
Env   : SUPABASE_URL, SUPABASE_SERVICE_KEY

params-json :
{
  "report": "commune" | "avant_apres" | "synthese",
  "communeId": "uuid (requis pour commune / avant_apres)",
  "siteIds": ["uuid", ...] (optionnel),
  "ids": ["uuid", ...] (optionnel — factures présélectionnées depuis Mes documents),
  "from": "YYYY-MM-DD", "to": "YYYY-MM-DD" (optionnels),
  "dataLogger": true|false
}

Principes :
- Séries temporelles chronologiques partout (X = semestres 2019-S1 → …), axes titrés avec unités.
- Périodes de facturation ventilées PRO-RATA JOURS sur les semestres calendaires.
- Avant/après travaux : dates RÉELLES SMEM par commune (mail du 03/07/2026), fenêtre de
  travaux EXCLUE, moyennes annualisées.
- Feuille « Données » masquée (source des TCD), triée chronologiquement → TCD ordonnés.
- Données simulées : disclaimers explicites sur la page de garde.
"""
import json
import math
import os
import sys
import urllib.request
import urllib.parse
import zipfile
import shutil
from pathlib import Path
from datetime import date, datetime, timedelta
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.drawing.fill import PatternFillProperties
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Styles Ability ────────────────────────────────────────────────────────────
ORANGE = "F97316"
ORANGE_DARK = "EA580C"
SOFT = "FFEDD5"
GREY = "585E74"
HDR_FILL = PatternFill("solid", fgColor=ORANGE)
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=15, color=ORANGE_DARK)
SUB_FONT = Font(italic=True, size=10, color=GREY)
KPI_FILL = PatternFill("solid", fgColor=SOFT)
EUR_FMT = '#,##0.00'
EUR0_FMT = '#,##0'
KWH_FMT = '#,##0.0'
PCT_FMT = '0.0'  # unité (%) portée par l'en-tête de colonne, pas par la cellule
SMEM_LOGO_PATH = Path(__file__).resolve().parents[2] / "logo_smem.png"

# ── Accès Supabase (REST, paginé) ────────────────────────────────────────────
SB_URL = os.environ.get("SUPABASE_URL", "").rstrip("/")
SB_KEY = os.environ.get("SUPABASE_SERVICE_KEY", "")

def sb_get(path: str, params: dict) -> list:
    rows, start = [], 0
    while True:
        qs = urllib.parse.urlencode(params)
        req = urllib.request.Request(f"{SB_URL}/rest/v1/{path}?{qs}")
        req.add_header("apikey", SB_KEY)
        req.add_header("Authorization", f"Bearer {SB_KEY}")
        req.add_header("Range", f"{start}-{start + 999}")
        with urllib.request.urlopen(req, timeout=60) as r:
            batch = json.loads(r.read().decode())
        rows.extend(batch)
        if len(batch) < 1000:
            return rows
        start += 1000

INVOICE_SELECT = "id,facture_number,facture_date,total_ht,tva,autres_taxes,total_ttc,categorie,commune_id,site_id,communes(nom),sites(nom,pdl,kva)"

def load_data(p: dict):
    base = {"select": INVOICE_SELECT, "archived": "eq.false", "order": "facture_date.asc"}
    if p.get("communeId"):
        base["commune_id"] = f"eq.{p['communeId']}"
    if p.get("from"):
        base["facture_date"] = f"gte.{p['from']}"
    ids = p.get("ids") or []
    if ids:
        invoices = []
        for i in range(0, len(ids), 60):  # limite de longueur d'URL PostgREST
            q = dict(base)
            q["id"] = f"in.({','.join(ids[i:i + 60])})"
            invoices.extend(sb_get("invoices", q))
    else:
        invoices = sb_get("invoices", base)
    if p.get("to"):
        invoices = [i for i in invoices if i["facture_date"] <= p["to"]]
    site_ids = set(p.get("siteIds") or [])
    if site_ids:
        invoices = [i for i in invoices if i.get("site_id") in site_ids]
    inv_ids = {i["id"] for i in invoices}
    periods = [r for r in sb_get("consumption_periods", {
        "select": "invoice_id,poste_tarifaire,period_start,period_end,consommation_kwh,prix_unitaire_ckwh,montant_eur"}) if r["invoice_id"] in inv_ids]
    charges = [r for r in sb_get("invoice_charges", {
        "select": "invoice_id,category,libelle,period_start,period_end,montant_eur"}) if r["invoice_id"] in inv_ids]
    communes = sb_get("communes", {"select": "id,nom,points_lumineux,armoires,travaux_debut,travaux_fin,travaux_estimes"})
    return invoices, periods, charges, {c["id"]: c for c in communes}

# ── Normalisation pro-rata (semestres calendaires) ───────────────────────────
def classify(poste: str) -> str:
    s = (poste or "").lower()
    if "pleine" in s or s.startswith("hp"):
        return "HP"
    if "creuse" in s or s.startswith("hc"):
        return "HC"
    return "Base"

def d(s):
    return date.fromisoformat(s) if s else None

def semester_buckets(start: date, end: date):
    out = []
    y = start.year
    while y <= end.year:
        for half, (b0, b1) in ((1, (date(y, 1, 1), date(y, 6, 30))), (2, (date(y, 7, 1), date(y, 12, 31)))):
            lo, hi = max(start, b0), min(end, b1)
            if lo <= hi:
                out.append((y, half, (hi - lo).days + 1))
        y += 1
    return out

def fallback_range(inv):
    fd = d(inv["facture_date"])
    return fd - timedelta(days=181), fd

def normalize(invoices, periods, charges):
    """→ lignes longues (facture × poste × bucket semestriel), pro-rata jours, TRIÉES chronologiquement."""
    inv_by_id = {i["id"]: i for i in invoices}
    period_range = {}
    for r in periods:
        if r["period_start"] and r["period_end"]:
            cur = period_range.get(r["invoice_id"])
            lo, hi = d(r["period_start"]), d(r["period_end"])
            period_range[r["invoice_id"]] = (min(lo, cur[0]) if cur else lo, max(hi, cur[1]) if cur else hi)
    rows = []

    def push(inv, type_, poste, start, end, kwh, eur):
        if not start or not end:
            start, end = fallback_range(inv)
        total_days = (end - start).days + 1
        for (yy, hh, days) in semester_buckets(start, end):
            f = days / total_days
            rows.append({
                "commune": (inv.get("communes") or {}).get("nom", "—"),
                "site": (inv.get("sites") or {}).get("nom", "—"),
                "cat": "Éclairage public" if inv.get("categorie") == "eclairage_public" else "Bâtiment",
                "type": type_, "poste": poste, "annee": yy, "sem": hh, "periode": f"{yy}-S{hh}",
                "kwh": kwh * f, "eur": eur * f, "days": days,
                "num": inv["facture_number"], "date": inv["facture_date"],
            })

    for r in periods:
        inv = inv_by_id.get(r["invoice_id"])
        if inv:
            push(inv, "Consommation", classify(r["poste_tarifaire"]), d(r["period_start"]), d(r["period_end"]),
                 float(r["consommation_kwh"] or 0), float(r["montant_eur"] or 0))
    for r in charges:
        inv = inv_by_id.get(r["invoice_id"])
        if not inv:
            continue
        typ = "Part fixe" if r["category"] == "fixed" else "Taxes"
        rng = period_range.get(r["invoice_id"])
        start = d(r["period_start"]) or (rng[0] if rng else None)
        end = d(r["period_end"]) or (rng[1] if rng else None)
        push(inv, typ, typ if typ == "Part fixe" else (r["libelle"] or "Taxe"), start, end, 0.0, float(r["montant_eur"] or 0))

    rows.sort(key=lambda r: (r["annee"], r["sem"], r["commune"], r["site"], r["type"], r["poste"]))
    return rows

def periods_sorted(rows):
    return sorted({r["periode"] for r in rows})

def trim_partial_semesters(rows):
    """Retire les semestres calendaires de bord partiellement couverts (artefacts de
    ventilation pro-rata : demi-hauteur en début/fin de série → chute trompeuse).
    Un semestre est conservé si sa couverture (somme des jours ventilés) atteint
    ≥ 85 % de la couverture médiane (intérieurs ~100 %, demi-semestres de bord ~50-80 %)."""
    if not rows:
        return rows
    dsum = defaultdict(float)
    for r in rows:
        dsum[r["periode"]] += r.get("days", 0)
    vals = sorted(dsum.values())
    med = vals[len(vals) // 2] if vals else 0
    if med <= 0:
        return rows
    keep = {p for p, v in dsum.items() if v >= 0.85 * med}
    return [r for r in rows if r["periode"] in keep]

# ── Avant/après : allocation par fenêtres de dates RÉELLES ───────────────────
TODAY = date.today()
BEFORE_START = date(2017, 1, 1)  # les données simulées démarrent en 2017

def overlap_days(a0, a1, b0, b1):
    lo, hi = max(a0, b0), min(a1, b1)
    return max(0, (hi - lo).days + 1)

def alloc_windows(invoices, periods, charges, debut: date, fin: date, keyf):
    """Ventile kWh (conso) et € (toutes composantes) sur AVANT [2019, debut) et APRÈS (fin, auj.],
    pro-rata jours ; la fenêtre de travaux est exclue. Rend {key: {kb,ka,eb,ea}} annualisés ensuite."""
    inv_by_id = {i["id"]: i for i in invoices}
    # Borne « après » = dernière période réellement facturée (et non aujourd'hui) : sinon
    # l'annualisation diviserait par une fenêtre plus large que les données → moyenne sous-estimée.
    last_end = max((d(p["period_end"]) for p in periods if p.get("period_end")), default=TODAY)
    before = (BEFORE_START, debut - timedelta(days=1))
    after = (fin + timedelta(days=1), min(TODAY, last_end))
    out = defaultdict(lambda: {"kb": 0.0, "ka": 0.0, "eb": 0.0, "ea": 0.0})

    def add(inv, start, end, kwh, eur):
        if not start or not end:
            start, end = fallback_range(inv)
        total = (end - start).days + 1
        k = keyf(inv)
        fb = overlap_days(start, end, *before) / total
        fa = overlap_days(start, end, *after) / total
        out[k]["kb"] += kwh * fb
        out[k]["ka"] += kwh * fa
        out[k]["eb"] += eur * fb
        out[k]["ea"] += eur * fa

    for r in periods:
        inv = inv_by_id.get(r["invoice_id"])
        if inv:
            add(inv, d(r["period_start"]), d(r["period_end"]), float(r["consommation_kwh"] or 0), float(r["montant_eur"] or 0))
    for r in charges:
        inv = inv_by_id.get(r["invoice_id"])
        if inv:
            add(inv, d(r["period_start"]), d(r["period_end"]), 0.0, float(r["montant_eur"] or 0))

    yb = max(0.25, ((before[1] - before[0]).days + 1) / 365.25)
    ya = max(0.25, ((after[1] - after[0]).days + 1) / 365.25)
    return {k: {"kwh_avant": v["kb"] / yb, "kwh_apres": v["ka"] / ya,
                "eur_avant": v["eb"] / yb, "eur_apres": v["ea"] / ya} for k, v in out.items()}

# ── Graphiques (séries temporelles, axes titrés) ─────────────────────────────
def make_banded_line_chart(ws, anchor, title, x_title, y_title, cats_ref, data_ref, band_ref=None):
    """Courbe temporelle ; bande de repère optionnelle via barres superposées."""
    line = LineChart()
    line.varyColors = False
    line.add_data(data_ref, titles_from_data=True)
    line.set_categories(cats_ref)  # catégories sur la COURBE (sinon axe 1,2,3 en graphe combiné)
    for cs in line.series:  # une seule couleur, trait plein (fini la courbe « divisée »)
        cs.graphicalProperties.line.solidFill = ORANGE
        cs.graphicalProperties.line.width = 28000  # EMU ≈ 2,2 pt
        cs.smooth = False
    if band_ref is not None:
        chart = BarChart()
        chart.type = "col"
        chart.gapWidth = 0
        chart.add_data(band_ref, titles_from_data=True)
        s = chart.series[0]
        s.graphicalProperties.solidFill = "D9D9D9"
        s.graphicalProperties.line.noFill = True
        chart.set_categories(cats_ref)
        chart += line  # la bande garde ses catégories, la courbe aussi
    else:
        chart = line
    chart.legend = None  # pas de label : la bande est expliquée par la note sous le graphe
    chart.title = title
    chart.height, chart.width = 9, 21
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.set_categories(cats_ref)
    ws.add_chart(chart, anchor)
    return chart

def make_timeseries_chart(ws, anchor, title, y_title, cats_ref, data_ref, band_ref=None):
    return make_banded_line_chart(ws, anchor, title, "Période (semestre)", y_title, cats_ref, data_ref, band_ref)

def make_bar_chart(ws, anchor, title, x_title, y_title, cats_ref, data_ref):
    ch = BarChart()
    ch.type = "col"
    ch.title = title
    ch.height, ch.width = 9, 21
    ch.x_axis.title = x_title
    ch.y_axis.title = y_title
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    ch.add_data(data_ref, titles_from_data=True)
    ch.set_categories(cats_ref)
    ws.add_chart(ch, anchor)
    return ch

def make_multi_line_chart(ws, anchor, title, x_title, y_title, cats_ref, data_refs):
    ch = LineChart()
    ch.title = title
    ch.height, ch.width = 9, 21
    ch.x_axis.title = x_title
    ch.y_axis.title = y_title
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    for ref in data_refs:
        ch.add_data(ref, titles_from_data=True)
    ch.set_categories(cats_ref)
    ws.add_chart(ch, anchor)
    return ch

def style_header_row(ws, row, upto):
    for c in range(1, upto + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(vertical="center")

def add_smem_logo(ws, anchor="D1", width=108):
    if not SMEM_LOGO_PATH.exists():
        return
    try:
        logo = XLImage(str(SMEM_LOGO_PATH))
    except Exception:
        return
    ratio = logo.height / logo.width if logo.width else 1
    logo.width = width
    logo.height = int(logo.width * ratio)
    ws.add_image(logo, anchor)

# ── Feuilles ─────────────────────────────────────────────────────────────────
def travaux_label(ci):
    if not ci or not ci.get("travaux_debut"):
        return "n.d."
    est = " (estimé)" if ci.get("travaux_estimes") else ""
    return f"{datetime.fromisoformat(ci['travaux_debut']).strftime('%d/%m/%Y')} → {datetime.fromisoformat(ci['travaux_fin']).strftime('%d/%m/%Y')}{est}"

def sheet_garde(wb, p, invoices, scope_label, ci=None):
    ws = wb.active
    ws.title = "Garde"
    ws.merge_cells("A1:D1")
    ws.merge_cells("A2:D2")
    add_smem_logo(ws, "F1", 140)
    ws["A1"] = "Rapport d'analyse des factures d'électricité"
    ws["A1"].font = Font(bold=True, size=18, color=ORANGE_DARK)
    ws["A1"].alignment = Alignment(vertical="center")
    ws["A2"] = "Ability x Syndicat Mixte d'Électricité de la Martinique (SMEM)"
    ws["A2"].font = Font(bold=True, size=11, color=GREY)
    ws["A2"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 24
    labels = {"commune": "Rapport par commune", "avant_apres": "Rapport avant / après travaux (PEPP)", "synthese": "Rapport de synthèse"}
    items = [
        ("Type de rapport", labels.get(p["report"], p["report"])),
        ("Périmètre", scope_label),
        ("Factures analysées", len(invoices)),
        ("Période couverte", f"{min((i['facture_date'] for i in invoices), default='—')} → {max((i['facture_date'] for i in invoices), default='—')}"),
        ("Généré le", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Données data logger", "Incluses (démonstration)" if p.get("dataLogger") else "Non incluses (espace réservé)"),
    ]
    if ci:
        items.insert(2, ("Travaux d'éclairage public", travaux_label(ci)))
        if ci.get("points_lumineux"):
            items.insert(3, ("Parc EP (référentiel SMEM)", f"{ci['points_lumineux']} points lumineux · {ci['armoires']} armoires"))
    ws["A4"] = "Synthèse du périmètre"
    style_header_row(ws, 4, 2)
    r = 5
    for k, v in items:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v).fill = KPI_FILL
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    note_row = r + 1
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row + 2, end_column=6)
    ws.cell(row=note_row, column=1, value=(
        "AVERTISSEMENT : rapport de démonstration. Les factures sont simulées mais calibrées sur des ordres de grandeur réels ; "
        "les dates de travaux proviennent du référentiel SMEM du 03/07/2026 et les communes non documentées restent marquées « estimé ». "
        "La feuille masquée « Données » contient le détail normalisé utilisé par les tableaux croisés dynamiques."
    ))
    ws.cell(row=note_row, column=1).font = SUB_FONT
    ws.cell(row=note_row, column=1).fill = PatternFill("solid", fgColor="FFF7ED")
    ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 4
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    return ws

DATA_HEADERS = ["Commune", "Site", "Catégorie", "Type", "Poste", "Année", "Semestre", "Période", "kWh", "Montant €", "Prix moyen c€/kWh", "N° facture", "Date facture"]
COL = {h: i for i, h in enumerate(DATA_HEADERS)}

def sheet_donnees(wb, rows):
    ws = wb.create_sheet("Données")
    ws.append(DATA_HEADERS)
    for r in rows:
        prix = round(r["eur"] / r["kwh"] * 100, 2) if r["kwh"] > 0 and r["type"] == "Consommation" else None
        ws.append([r["commune"], r["site"], r["cat"], r["type"], r["poste"], r["annee"], f"S{r['sem']}", r["periode"],
                   round(r["kwh"], 1), round(r["eur"], 2), prix, r["num"], r["date"]])
    style_header_row(ws, 1, len(DATA_HEADERS))
    ws.sheet_state = "hidden"  # source des TCD, volontairement masquée
    return ws

def in_works_window(periode: str, ci) -> bool:
    if not ci or not ci.get("travaux_debut"):
        return False
    y, s = periode.split("-S")
    sem_start = date(int(y), 1 if s == "1" else 7, 1)
    sem_end = date(int(y), 6 if s == "1" else 12, 30)
    return sem_start <= d(ci["travaux_fin"]) and sem_end >= d(ci["travaux_debut"])

def sheet_evolution(wb, rows, ci=None, title_suffix=""):
    """Séries temporelles kWh & € de l'ÉCLAIRAGE PUBLIC (objet des travaux : la bande
    grisée marque la fenêtre de rénovation) + décomposition annuelle compacte."""
    ws = wb.create_sheet("Évolution")
    # Courbes centrées sur l'éclairage public : c'est le poste rénové, seul à refléter les
    # travaux ; les bâtiments (peu impactés) diluaient la baisse. Repli sur tout si pas d'EP.
    ep_rows = [r for r in rows if r["cat"] == "Éclairage public"]
    rows = ep_rows or rows
    ws["A1"] = f"Évolution semestrielle — éclairage public{title_suffix}"
    ws["A1"].font = Font(bold=True, size=13, color=ORANGE_DARK)
    ws["A2"] = "Éclairage public uniquement (poste rénové). Montants toutes composantes ; kWh = énergie consommée. Bâtiments détaillés dans « Par site » / TCD."
    ws["A2"].font = SUB_FONT

    pers = periods_sorted(rows)
    kwh_by, eur_by = defaultdict(float), defaultdict(float)
    for r in rows:
        eur_by[r["periode"]] += r["eur"]
        if r["type"] == "Consommation":
            kwh_by[r["periode"]] += r["kwh"]
    kmax = max([kwh_by[p] for p in pers] or [0])
    emax = max([eur_by[p] for p in pers] or [0])

    headers = ["Période", "Consommation (kWh)", "Dépense (€)"]
    with_band = ci is not None and ci.get("travaux_debut")
    hr = 4
    for j, h in enumerate(headers):
        ws.cell(row=hr, column=1 + j, value=h)
    style_header_row(ws, hr, len(headers))
    ri = hr + 1
    for p in pers:
        ws.append([p, round(kwh_by[p]), round(eur_by[p], 2)])
        ws.cell(row=ri, column=2).number_format = KWH_FMT
        ws.cell(row=ri, column=3).number_format = EUR_FMT
        ri += 1
    last = ri - 1
    cats = Reference(ws, min_col=1, min_row=hr + 1, max_row=last)
    kwh_ref = Reference(ws, min_col=2, min_row=hr, max_row=last)
    eur_ref = Reference(ws, min_col=3, min_row=hr, max_row=last)
    if with_band:
        ws.cell(row=hr, column=6, value="Repère travaux kWh")
        ws.cell(row=hr, column=7, value="Repère travaux €")
        for row_idx, period in enumerate(pers, start=hr + 1):
            band = in_works_window(period, ci)
            ws.cell(row=row_idx, column=6, value=round(kmax * 1.05) if band else None)
            ws.cell(row=row_idx, column=7, value=round(emax * 1.05) if band else None)
        ws.column_dimensions["F"].hidden = True
        ws.column_dimensions["G"].hidden = True
    band_k = Reference(ws, min_col=6, min_row=hr, max_row=last) if with_band else None
    band_e = Reference(ws, min_col=7, min_row=hr, max_row=last) if with_band else None
    anchor_row = last + 2
    make_timeseries_chart(ws, f"A{anchor_row}", f"Consommation par semestre{title_suffix}", "Consommation (kWh)", cats, kwh_ref, band_k)
    make_timeseries_chart(ws, f"A{anchor_row + 19}", f"Dépense par semestre{title_suffix}", "Dépense (€ TTC)", cats, eur_ref, band_e)
    if with_band:
        ws.cell(row=anchor_row + 38, column=1,
                value=f"Bande grisée = fenêtre de travaux d'éclairage public : {travaux_label(ci)}.").font = SUB_FONT

    # Décomposition tarifaire compacte (par année, en €)
    years = sorted({r["annee"] for r in rows})
    postes = ["Base", "HP", "HC", "Part fixe", "Taxes"]
    def poste_of(r):
        return r["poste"] if r["type"] == "Consommation" else ("Part fixe" if r["type"] == "Part fixe" else "Taxes")
    dec = defaultdict(float)
    for r in rows:
        dec[(poste_of(r), r["annee"])] += r["eur"]
    top = anchor_row + 40
    ws.cell(row=top, column=1, value="Décomposition de la dépense par composante tarifaire (€)").font = Font(bold=True, size=12, color=ORANGE_DARK)
    ws.cell(row=top + 1, column=1, value="Composante")
    for j, y in enumerate(years):
        ws.cell(row=top + 1, column=2 + j, value=y)
    style_header_row(ws, top + 1, 1 + len(years))
    for i, po in enumerate(postes):
        ws.cell(row=top + 2 + i, column=1, value=po)
        for j, y in enumerate(years):
            c = ws.cell(row=top + 2 + i, column=2 + j, value=round(dec.get((po, y), 0), 2))
            c.number_format = EUR0_FMT
    ws.column_dimensions["A"].width = 24
    for col in "BCDE":
        ws.column_dimensions[col].width = 16
    return ws

def write_matrix(ws, top, left_label, col_keys, line_keys, values, fmt, title):
    ws.cell(row=top, column=1, value=title).font = Font(bold=True, size=12, color=ORANGE_DARK)
    hr = top + 1
    ws.cell(row=hr, column=1, value=left_label)
    for j, ck in enumerate(col_keys):
        ws.cell(row=hr, column=2 + j, value=ck)
    style_header_row(ws, hr, len(col_keys) + 1)
    for i, lk in enumerate(line_keys):
        ws.cell(row=hr + 1 + i, column=1, value=lk)
        for j, ck in enumerate(col_keys):
            v = values.get((lk, ck))
            cell = ws.cell(row=hr + 1 + i, column=2 + j, value=round(v, 2) if v else None)
            cell.number_format = fmt
    tr = hr + 1 + len(line_keys)
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
    for j, ck in enumerate(col_keys):
        cell = ws.cell(row=tr, column=2 + j, value=round(sum(values.get((lk, ck)) or 0 for lk in line_keys), 2))
        cell.number_format = fmt
        cell.font = Font(bold=True)
        cell.fill = KPI_FILL
    ws.column_dimensions["A"].width = 26
    return tr

def sheet_par_site(wb, rows):
    ws = wb.create_sheet("Par site")
    pers = periods_sorted(rows)
    sites = sorted({r["site"] for r in rows})
    eur = defaultdict(float)
    kwh = defaultdict(float)
    for r in rows:
        eur[(r["site"], r["periode"])] += r["eur"]
        if r["type"] == "Consommation":
            kwh[(r["site"], r["periode"])] += r["kwh"]
    tr1 = write_matrix(ws, 1, "Site", pers, sites, eur, EUR0_FMT, "Dépense par site et par semestre (€, toutes composantes)")
    write_matrix(ws, tr1 + 3, "Site", pers, sites, kwh, KWH_FMT, "Consommation par site et par semestre (kWh)")
    return ws

def sheet_avant_apres(wb, rows, alloc_site, ci, commune_nom):
    """UNE feuille : méthode, tableau par site, 2 histogrammes (X = sites), 2 courbes avec bande travaux."""
    ws = wb.create_sheet("Avant-Après")
    ws["A1"] = f"Avant / après travaux d'éclairage public — {commune_nom}"
    ws["A1"].font = Font(bold=True, size=13, color=ORANGE_DARK)
    meth = [
        f"Travaux : {travaux_label(ci)} (référentiel SMEM).",
        "Méthode : AVANT = début des données (2017) → veille du lancement ; APRÈS = lendemain de l'achèvement → dernière facture.",
        "La fenêtre de travaux est EXCLUE. Moyennes annualisées (pro-rata jours). € = toutes composantes.",
    ]
    for i, m in enumerate(meth):
        ws.cell(row=2 + i, column=1, value=m).font = SUB_FONT

    headers = ["Site", "kWh/an avant", "kWh/an après", "Δ conso (%)", "€/an avant", "€/an après", "Δ dépense (%)"]
    hr = 6
    for j, h in enumerate(headers):
        ws.cell(row=hr, column=1 + j, value=h)
    style_header_row(ws, hr, len(headers))
    sites = sorted(alloc_site.keys())
    ri = hr + 1
    for s in sites:
        a = alloc_site[s]
        dk = (a["kwh_apres"] - a["kwh_avant"]) / a["kwh_avant"] * 100 if a["kwh_avant"] else None
        de = (a["eur_apres"] - a["eur_avant"]) / a["eur_avant"] * 100 if a["eur_avant"] else None
        ws.append([s, round(a["kwh_avant"]), round(a["kwh_apres"]), dk, round(a["eur_avant"], 2), round(a["eur_apres"], 2), de])
        for col, fmt in ((2, KWH_FMT), (3, KWH_FMT), (4, PCT_FMT), (5, EUR0_FMT), (6, EUR0_FMT), (7, PCT_FMT)):
            ws.cell(row=ri, column=col).number_format = fmt
        ri += 1
    # TOTAL
    tk_av = sum(alloc_site[s]["kwh_avant"] for s in sites)
    tk_ap = sum(alloc_site[s]["kwh_apres"] for s in sites)
    te_av = sum(alloc_site[s]["eur_avant"] for s in sites)
    te_ap = sum(alloc_site[s]["eur_apres"] for s in sites)
    ws.append(["TOTAL", round(tk_av), round(tk_ap), (tk_ap - tk_av) / tk_av * 100 if tk_av else None,
               round(te_av, 2), round(te_ap, 2), (te_ap - te_av) / te_av * 100 if te_av else None])
    for col, fmt in ((2, KWH_FMT), (3, KWH_FMT), (4, PCT_FMT), (5, EUR0_FMT), (6, EUR0_FMT), (7, PCT_FMT)):
        cell = ws.cell(row=ri, column=col)
        cell.number_format = fmt
        cell.font = Font(bold=True)
        cell.fill = KPI_FILL
    ws.cell(row=ri, column=1).font = Font(bold=True)
    last_site_row = ri - 1

    # Histogrammes X = sites
    cats = Reference(ws, min_col=1, min_row=hr + 1, max_row=last_site_row)
    a_row = ri + 2
    make_bar_chart(ws, f"A{a_row}", "Consommation moyenne annuelle par site — avant vs après travaux",
                   "Sites de la commune", "Consommation (kWh / an)",
                   cats, Reference(ws, min_col=2, max_col=3, min_row=hr, max_row=last_site_row))
    make_bar_chart(ws, f"A{a_row + 19}", "Dépense moyenne annuelle par site — avant vs après travaux",
                   "Sites de la commune", "Dépense (€ / an)",
                   cats, Reference(ws, min_col=5, max_col=6, min_row=hr, max_row=last_site_row))

    # Courbes temporelles de la commune avec bande travaux (données à droite du tableau)
    pers = periods_sorted(rows)
    kwh_by, eur_by = defaultdict(float), defaultdict(float)
    for r in rows:
        eur_by[r["periode"]] += r["eur"]
        if r["type"] == "Consommation":
            kwh_by[r["periode"]] += r["kwh"]
    kmax = max([kwh_by[p] for p in pers] or [0])
    emax = max([eur_by[p] for p in pers] or [0])
    base_col = 10  # colonne J : bloc série temporelle
    ws.cell(row=hr - 1, column=base_col, value="Série temporelle (source des courbes)").font = SUB_FONT
    for j, h in enumerate(["Période", "kWh", "€", "Travaux (kWh)", "Travaux (€)"]):
        ws.cell(row=hr, column=base_col + j, value=h)
    style_header_row_range(ws, hr, base_col, base_col + 4)
    for i, p in enumerate(pers):
        band = in_works_window(p, ci)
        ws.cell(row=hr + 1 + i, column=base_col, value=p)
        ws.cell(row=hr + 1 + i, column=base_col + 1, value=round(kwh_by[p]))
        ws.cell(row=hr + 1 + i, column=base_col + 2, value=round(eur_by[p], 2))
        ws.cell(row=hr + 1 + i, column=base_col + 3, value=round(kmax * 1.05) if band else None)
        ws.cell(row=hr + 1 + i, column=base_col + 4, value=round(emax * 1.05) if band else None)
    lastp = hr + len(pers)
    cats_t = Reference(ws, min_col=base_col, min_row=hr + 1, max_row=lastp)
    make_timeseries_chart(ws, f"J{a_row}", f"Consommation semestrielle — {commune_nom}", "Consommation (kWh)",
                          cats_t, Reference(ws, min_col=base_col + 1, min_row=hr, max_row=lastp),
                          Reference(ws, min_col=base_col + 3, min_row=hr, max_row=lastp))
    make_timeseries_chart(ws, f"J{a_row + 19}", f"Dépense semestrielle — {commune_nom}", "Dépense (€ TTC)",
                          cats_t, Reference(ws, min_col=base_col + 2, min_row=hr, max_row=lastp),
                          Reference(ws, min_col=base_col + 4, min_row=hr, max_row=lastp))
    ws.column_dimensions["A"].width = 26
    return ws

def style_header_row_range(ws, row, c0, c1):
    for c in range(c0, c1 + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT

def sheet_synthese_avant_apres(wb, invoices, periods, charges, communes_info):
    """Synthèse : avant/après PAR COMMUNE (dates réelles propres à chacune) + histogramme X = communes."""
    ws = wb.create_sheet("Avant-Après communes")
    ws["A1"] = "Avant / après travaux — comparaison par commune (dates réelles SMEM)"
    ws["A1"].font = Font(bold=True, size=13, color=ORANGE_DARK)
    ws["A2"] = "AVANT = début des données (2017) → lancement ; APRÈS = achèvement → dernière facture ; fenêtre de travaux exclue ; moyennes annualisées."
    ws["A2"].font = SUB_FONT
    headers = ["Commune", "Travaux (SMEM)", "Points lum.", "Armoires", "kWh/an avant", "kWh/an après", "Δ conso (%)", "€/an avant", "€/an après", "Δ dépense (%)"]
    hr = 4
    for j, h in enumerate(headers):
        ws.cell(row=hr, column=1 + j, value=h)
    style_header_row(ws, hr, len(headers))
    by_commune = defaultdict(list)
    for i in invoices:
        by_commune[i.get("commune_id")].append(i)
    ri = hr + 1
    for cid, invs in sorted(by_commune.items(), key=lambda kv: (kv[1][0].get("communes") or {}).get("nom", "")):
        ci = communes_info.get(cid)
        nom = (invs[0].get("communes") or {}).get("nom", "—")
        if not ci or not ci.get("travaux_debut"):
            ws.append([nom, "n.d."] + [None] * 8)
            ri += 1
            continue
        ids = {i["id"] for i in invs}
        a = alloc_windows(invs, [p for p in periods if p["invoice_id"] in ids],
                          [c for c in charges if c["invoice_id"] in ids],
                          d(ci["travaux_debut"]), d(ci["travaux_fin"]), lambda inv: nom)
        v = a.get(nom, {"kwh_avant": 0, "kwh_apres": 0, "eur_avant": 0, "eur_apres": 0})
        dk = (v["kwh_apres"] - v["kwh_avant"]) / v["kwh_avant"] * 100 if v["kwh_avant"] else None
        de = (v["eur_apres"] - v["eur_avant"]) / v["eur_avant"] * 100 if v["eur_avant"] else None
        ws.append([nom, travaux_label(ci), ci.get("points_lumineux"), ci.get("armoires"),
                   round(v["kwh_avant"]), round(v["kwh_apres"]), dk, round(v["eur_avant"], 2), round(v["eur_apres"], 2), de])
        for col, fmt in ((5, KWH_FMT), (6, KWH_FMT), (7, PCT_FMT), (8, EUR0_FMT), (9, EUR0_FMT), (10, PCT_FMT)):
            ws.cell(row=ri, column=col).number_format = fmt
        ri += 1
    make_bar_chart(ws, f"A{ri + 2}", "Consommation moyenne annuelle par commune — avant vs après travaux",
                   "Communes", "Consommation (kWh / an)",
                   Reference(ws, min_col=1, min_row=hr + 1, max_row=ri - 1),
                   Reference(ws, min_col=5, max_col=6, min_row=hr, max_row=ri - 1))
    for i, w in enumerate([20, 30, 10, 9, 14, 14, 10, 13, 13, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws

def semester_range(year: int, sem: int):
    return (date(year, 1, 1), date(year, 6, 30)) if sem == 1 else (date(year, 7, 1), date(year, 12, 31))

def each_day(start: date, end: date):
    cur = start
    while cur <= end:
        yield cur
        cur += timedelta(days=1)

def pick_datalogger_example(invoices, rows):
    site_year = defaultdict(lambda: {"s1": 0.0, "s2": 0.0, "total": 0.0, "sems": set()})
    site_kva = {}
    for inv in invoices:
        site = (inv.get("sites") or {}).get("nom") or "—"
        kva = (inv.get("sites") or {}).get("kva")
        if kva not in (None, ""):
            try:
                site_kva[site] = max(float(kva), site_kva.get(site, 0))
            except Exception:
                pass
    for r in rows:
        if r["type"] != "Consommation" or not r["kwh"]:
            continue
        bucket = site_year[(r["site"], r["annee"])]
        bucket[f"s{r['sem']}"] += r["kwh"]
        bucket["total"] += r["kwh"]
        bucket["sems"].add(r["sem"])
    if not site_year:
        return None
    (site, year), bucket = max(site_year.items(), key=lambda item: (item[0][1], len(item[1]["sems"]), item[1]["total"]))
    return {
        "site": site,
        "year": year,
        "invoice_sem": {1: bucket["s1"], 2: bucket["s2"]},
        "subscribed_kva": site_kva.get(site) or 36.0,
    }

def simulate_logger_targets(invoice_sem):
    factors = {1: 0.991, 2: 1.028}
    return {sem: round((invoice_sem.get(sem) or 0.0) * factors[sem], 1) for sem in (1, 2)}

def simulate_logger_daily(year, logger_sem, subscribed_kva):
    daily_energy = []
    for sem in (1, 2):
        start, end = semester_range(year, sem)
        days = list(each_day(start, end))
        weights = []
        for idx, day in enumerate(days):
            season = 1 + 0.11 * math.sin(((day.timetuple().tm_yday - 1) / 365) * 2 * math.pi)
            weekly = 1 + 0.04 * math.sin(idx * 2 * math.pi / 7)
            pulse = 1.06 if day.day in (5, 14, 23) else 1.0
            weights.append(max(0.35, season * weekly * pulse))
        total_weight = sum(weights) or 1.0
        scale = (logger_sem.get(sem) or 0.0) / total_weight
        for day, weight in zip(days, weights):
            daily_energy.append((day, weight * scale))
    mean_daily = (sum(v for _, v in daily_energy) / len(daily_energy)) if daily_energy else 0.0
    daily_power = []
    for idx, (day, kwh) in enumerate(daily_energy):
        ratio = (kwh / mean_daily) if mean_daily else 1.0
        peak = subscribed_kva * (0.67 + 0.18 * ratio + 0.05 * math.sin(idx * 0.17) + (0.09 if idx % 37 == 0 else 0.0))
        peak = max(subscribed_kva * 0.4, min(subscribed_kva * 1.12, peak))
        daily_power.append((day, round(peak, 2)))
    return daily_energy, daily_power

def aggregate_weekly(daily_energy):
    weekly = defaultdict(float)
    for day, kwh in daily_energy:
        week_start = day - timedelta(days=day.weekday())
        weekly[week_start] += kwh
    out = []
    for week_start in sorted(weekly):
        marker_day = week_start + timedelta(days=3)
        sem = 1 if marker_day <= date(marker_day.year, 6, 30) else 2
        out.append((week_start, round(weekly[week_start], 1), sem))
    return out

def sheet_datalogger(wb, p, scope_label, invoices, rows):
    if not p.get("dataLogger"):
        ws = wb.create_sheet("Data loggers")
        ws["A1"] = "Données complémentaires — data loggers d'armoires"
        ws["A1"].font = Font(bold=True, size=13, color=ORANGE_DARK)
        ws["A3"] = "EMPLACEMENT RÉSERVÉ — le connecteur data logger n'est pas encore raccordé."
        ws["A3"].font = Font(bold=True, color="B45309")
        ws["A3"].fill = KPI_FILL
        ws["A5"] = "À la connexion : consommation quotidienne réelle et historique de coupures, indépendants des factures."
        ws["A5"].font = SUB_FONT
        ws.column_dimensions["A"].width = 100
        return ws

    example = pick_datalogger_example(invoices, rows)
    ws = wb.create_sheet("Data logger (démo)")
    ws["A1"] = "Data logger — analyse agrégée (données fictives de démonstration)"
    ws["A1"].font = Font(bold=True, size=13, color="B91C1C")
    ws["A2"] = f"Données horaires / infra-horaires simulées puis agrégées à la semaine et au jour · périmètre : {scope_label}"
    ws["A2"].font = SUB_FONT

    if not example:
        ws["A4"] = "Aucune donnée de consommation disponible pour produire la démonstration data logger."
        ws["A4"].font = Font(bold=True, color="B45309")
        ws["A4"].fill = KPI_FILL
        ws.column_dimensions["A"].width = 100
        return ws

    site = example["site"]
    year = example["year"]
    subscribed_kva = float(example["subscribed_kva"] or 36.0)
    invoice_sem = example["invoice_sem"]
    logger_sem = simulate_logger_targets(invoice_sem)
    daily_energy, daily_power = simulate_logger_daily(year, logger_sem, subscribed_kva)
    weekly_energy = aggregate_weekly(daily_energy)

    ws["A4"] = "Site exemple"
    ws["B4"] = f"{site} · année {year} · abonnement {subscribed_kva:.0f} kVA"
    ws["A5"] = "Le logger fictif représente une mesure horaire / infra-horaire agrégée pour éviter les milliers de lignes brutes."
    ws["A5"].font = SUB_FONT

    ws["A7"] = "Tableau 1 — Synthèse comparative par période"
    ws["A7"].font = Font(bold=True, size=12, color=ORANGE_DARK)
    headers = ["Période", "Consommation Facture (kWh)", "Consommation Logger (kWh)", "Écart Absolu (kWh)", "Écart (%)", "Statut"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=8, column=j, value=h)
    style_header_row(ws, 8, len(headers))

    summary_rows = [
        ("Semestre 1", invoice_sem.get(1, 0.0), logger_sem.get(1, 0.0)),
        ("Semestre 2", invoice_sem.get(2, 0.0), logger_sem.get(2, 0.0)),
        ("Total annuel", sum(invoice_sem.values()), sum(logger_sem.values())),
    ]
    for idx, (label, facture_kwh, logger_kwh) in enumerate(summary_rows, start=9):
        ecart_abs = abs(logger_kwh - facture_kwh)
        ecart_pct = (ecart_abs / facture_kwh * 100) if facture_kwh else None
        statut = "Alerte" if (ecart_pct or 0) > 2 else "Conforme"
        ws.cell(row=idx, column=1, value=label)
        ws.cell(row=idx, column=2, value=round(facture_kwh, 1)).number_format = KWH_FMT
        ws.cell(row=idx, column=3, value=round(logger_kwh, 1)).number_format = KWH_FMT
        ws.cell(row=idx, column=4, value=round(ecart_abs, 1)).number_format = KWH_FMT
        ws.cell(row=idx, column=5, value=ecart_pct).number_format = PCT_FMT
        ws.cell(row=idx, column=6, value=statut)
        if statut == "Alerte":
            ws.cell(row=idx, column=6).fill = PatternFill("solid", fgColor="FECACA")
        else:
            ws.cell(row=idx, column=6).fill = PatternFill("solid", fgColor="DCFCE7")

    src = wb.create_sheet("Data logger source")
    src.sheet_state = "hidden"

    src["A1"] = "Semaine"
    src["B1"] = "Énergie logger (kWh/semaine)"
    src["C1"] = "Repère Semestre 2"
    style_header_row(src, 1, 3)
    max_weekly = max([v for _, v, _ in weekly_energy] or [0])
    for row_idx, (week_start, weekly_kwh, sem) in enumerate(weekly_energy, start=2):
        src.cell(row=row_idx, column=1, value=week_start.strftime("%d/%m/%Y"))
        src.cell(row=row_idx, column=2, value=weekly_kwh)
        src.cell(row=row_idx, column=3, value=round(max_weekly * 1.05, 1) if sem == 2 else None)

    src["E1"] = "Date"
    src["F1"] = "Puissance max quotidienne (kVA)"
    src["G1"] = "Abonnement (kVA)"
    style_header_row(src, 1, 7)
    for row_idx, ((day, peak), (_, _kwh)) in enumerate(zip(daily_power, daily_energy), start=2):
        src.cell(row=row_idx, column=5, value=day.strftime("%d/%m/%Y"))
        src.cell(row=row_idx, column=6, value=peak)
        src.cell(row=row_idx, column=7, value=round(subscribed_kva, 2))

    ws["A14"] = "Graphique 1 — Courbe de charge globale et périodes de facturation"
    ws["A14"].font = Font(bold=True, size=12, color=ORANGE_DARK)
    ws["A33"] = "Graphique 2 — Histogramme comparatif Facture vs Logger"
    ws["A33"].font = Font(bold=True, size=12, color=ORANGE_DARK)
    ws["A52"] = "Graphique 3 — Analyse des pics de puissance vs abonnement"
    ws["A52"].font = Font(bold=True, size=12, color=ORANGE_DARK)

    weekly_last = 1 + len(weekly_energy)
    make_banded_line_chart(
        ws, "A15",
        f"Courbe de charge globale agrégée à la semaine — {site} ({year})",
        "Semaine",
        "Énergie logger (kWh / semaine)",
        Reference(src, min_col=1, min_row=2, max_row=weekly_last),
        Reference(src, min_col=2, min_row=1, max_row=weekly_last),
        Reference(src, min_col=3, min_row=1, max_row=weekly_last),
    )
    ws["A31"] = "Bande grisée = Semestre 2 ; la série hebdomadaire agrège la mesure horaire / infra-horaire du logger."
    ws["A31"].font = SUB_FONT

    make_bar_chart(
        ws, "A34",
        f"Facture vs logger — {site} ({year})",
        "Période",
        "Consommation (kWh)",
        Reference(ws, min_col=1, min_row=9, max_row=11),
        Reference(ws, min_col=2, max_col=3, min_row=8, max_row=11),
    )

    daily_last = 1 + len(daily_power)
    make_multi_line_chart(
        ws, "A53",
        f"Pics de puissance quotidiens vs abonnement — {site}",
        "Jour",
        "Puissance (kVA)",
        Reference(src, min_col=5, min_row=2, max_row=daily_last),
        [
            Reference(src, min_col=6, min_row=1, max_row=daily_last),
            Reference(src, min_col=7, min_row=1, max_row=daily_last),
        ],
    )

    for i, w in enumerate([18, 24, 24, 18, 12, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws

def make_timeseries_chart_daily(ws, anchor, title, cats_ref, data_ref):
    ch = LineChart()
    ch.title = title
    ch.height, ch.width = 9, 22
    ch.x_axis.title = "Jour"
    ch.y_axis.title = "Énergie (kWh / jour)"
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    ch.add_data(data_ref, titles_from_data=True)
    ch.set_categories(cats_ref)
    ws.add_chart(ch, anchor)

# ── TCD natif (injection XML pivotCache refreshOnLoad) ───────────────────────
NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

def _cache_definition_xml(n_rows: int) -> bytes:
    fields = "".join(f'<cacheField name="{h}" numFmtId="0"><sharedItems containsBlank="1"/></cacheField>' for h in DATA_HEADERS)
    return (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<pivotCacheDefinition xmlns="{NS_MAIN}" xmlns:r="{NS_R}" r:id="rId1" refreshOnLoad="1" refreshedBy="Ability" '
        f'refreshedDate="45000" createdVersion="6" refreshedVersion="6" minRefreshableVersion="3" recordCount="0">'
        f'<cacheSource type="worksheet"><worksheetSource ref="A1:M{n_rows}" sheet="Données"/></cacheSource>'
        f'<cacheFields count="{len(DATA_HEADERS)}">{fields}</cacheFields>'
        f"</pivotCacheDefinition>"
    ).encode()

def _cache_records_xml() -> bytes:
    return (f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<pivotCacheRecords xmlns="{NS_MAIN}" xmlns:r="{NS_R}" count="0"/>').encode()

def _pivot_table_xml(name, cache_id, row_fields, col_fields, data_fields, ref) -> bytes:
    """data_fields : liste [(index_colonne, libellé)] — 2 valeurs max (Σ € et Σ kWh)."""
    k = len(DATA_HEADERS)
    data_idx = {i for i, _ in data_fields}
    pf = []
    for i in range(k):
        if i in row_fields:
            pf.append('<pivotField axis="axisRow" showAll="0"><items count="1"><item t="default"/></items></pivotField>')
        elif i in col_fields:
            pf.append('<pivotField axis="axisCol" showAll="0"><items count="1"><item t="default"/></items></pivotField>')
        elif i in data_idx:
            pf.append('<pivotField dataField="1" showAll="0"/>')
        else:
            pf.append('<pivotField showAll="0"/>')
    rows_xml = "".join(f'<field x="{i}"/>' for i in row_fields)
    cols_xml = "".join(f'<field x="{i}"/>' for i in col_fields)
    multi = len(data_fields) > 1
    if multi:
        cols_xml += '<field x="-2"/>'  # champ « Σ Valeurs »
    col_count = len(col_fields) + (1 if multi else 0)
    col_items = '<colItems count="2"><i><x/></i><i i="1"><x v="1"/></i></colItems>' if multi else '<colItems count="1"><i t="grand"><x/></i></colItems>'
    dfs = "".join(f'<dataField name="{label}" fld="{idx}" baseField="0" baseItem="0"/>' for idx, label in data_fields)
    return (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<pivotTableDefinition xmlns="{NS_MAIN}" name="{name}" cacheId="{cache_id}" dataOnRows="0" applyNumberFormats="0" '
        f'applyBorderFormats="0" applyFontFormats="0" applyPatternFormats="0" applyAlignmentFormats="0" '
        f'applyWidthHeightFormats="1" dataCaption="Valeurs" updatedVersion="6" createdVersion="6" minRefreshableVersion="3" '
        f'useAutoFormatting="1" itemPrintTitles="1" indent="0" outline="1" outlineData="1" multipleFieldFilters="0">'
        f'<location ref="{ref}" firstHeaderRow="1" firstDataRow="2" firstDataCol="1"/>'
        f'<pivotFields count="{k}">{"".join(pf)}</pivotFields>'
        f'<rowFields count="{len(row_fields)}">{rows_xml}</rowFields>'
        f'<rowItems count="1"><i t="grand"><x/></i></rowItems>'
        f'<colFields count="{col_count}">{cols_xml}</colFields>{col_items}'
        f'<dataFields count="{len(data_fields)}">{dfs}</dataFields>'
        f'<pivotTableStyleInfo name="PivotStyleMedium9" showRowHeaders="1" showColHeaders="1" showRowStripes="0" showColStripes="0" showLastColumn="1"/>'
        f"</pivotTableDefinition>"
    ).encode()

def inject_pivots(xlsx_path: str, n_data_rows: int, pivots: list):
    tmp = xlsx_path + ".tmp"
    with zipfile.ZipFile(xlsx_path, "r") as zin:
        contents = {n: zin.read(n) for n in zin.namelist()}
    wb_xml = contents["xl/workbook.xml"].decode()
    rels_xml = contents["xl/_rels/workbook.xml.rels"].decode()
    contents["xl/pivotCache/pivotCacheDefinition1.xml"] = _cache_definition_xml(n_data_rows)
    contents["xl/pivotCache/pivotCacheRecords1.xml"] = _cache_records_xml()
    contents["xl/pivotCache/_rels/pivotCacheDefinition1.xml.rels"] = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheRecords" Target="pivotCacheRecords1.xml"/>'
        "</Relationships>"
    ).encode()
    import re as _re
    next_rid = max(int(m) for m in _re.findall(r'Id="rId(\d+)"', rels_xml)) + 1
    cache_rid = f"rId{next_rid}"
    rels_xml = rels_xml.replace("</Relationships>",
        f'<Relationship Id="{cache_rid}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheDefinition" Target="pivotCache/pivotCacheDefinition1.xml"/></Relationships>')
    if "<pivotCaches>" not in wb_xml:
        wb_xml = wb_xml.replace("</workbook>",
            f'<pivotCaches><pivotCache cacheId="10" xmlns:r="{NS_R}" r:id="{cache_rid}"/></pivotCaches></workbook>')
    contents["xl/workbook.xml"] = wb_xml.encode()
    contents["xl/_rels/workbook.xml.rels"] = rels_xml.encode()
    ct = contents["[Content_Types].xml"].decode()
    for idx, p in enumerate(pivots, start=1):
        contents[f"xl/pivotTables/pivotTable{idx}.xml"] = _pivot_table_xml(p["name"], 10, p["rows"], p["cols"], p["data"], p["ref"])
        contents[f"xl/pivotTables/_rels/pivotTable{idx}.xml.rels"] = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheDefinition" Target="../pivotCache/pivotCacheDefinition1.xml"/>'
            "</Relationships>"
        ).encode()
        rels_path = f"xl/worksheets/_rels/sheet{p['sheet_index']}.xml.rels"
        rels = contents.get(rels_path,
            b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"></Relationships>').decode()
        rels = rels.replace("</Relationships>",
            f'<Relationship Id="rIdP{idx}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotTable" Target="../pivotTables/pivotTable{idx}.xml"/></Relationships>')
        contents[rels_path] = rels.encode()
        ct = ct.replace("</Types>",
            f'<Override PartName="/xl/pivotTables/pivotTable{idx}.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.pivotTable+xml"/></Types>')
    ct = ct.replace("</Types>",
        '<Override PartName="/xl/pivotCache/pivotCacheDefinition1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheDefinition+xml"/>'
        '<Override PartName="/xl/pivotCache/pivotCacheRecords1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheRecords+xml"/></Types>')
    contents["[Content_Types].xml"] = ct.encode()
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for n, data in contents.items():
            zout.writestr(n, data)
    shutil.move(tmp, xlsx_path)

# ── Assemblage ────────────────────────────────────────────────────────────────
def build(p: dict, out_path: str):
    invoices, periods, charges, communes_info = load_data(p)
    if not invoices:
        raise SystemExit("Aucune facture dans le périmètre demandé.")
    rows = trim_partial_semesters(normalize(invoices, periods, charges))

    ci = communes_info.get(p.get("communeId")) if p.get("communeId") else None
    if p["report"] in ("commune", "avant_apres"):
        scope = (ci or {}).get("nom") or rows[0]["commune"]
    else:
        scope = "Rapport général — toutes les communes"
    if p.get("ids"):
        scope += f" · {len(invoices)} factures sélectionnées manuellement"

    wb = Workbook()
    sheet_garde(wb, p, invoices, scope, ci if p["report"] in ("commune", "avant_apres") else None)

    if p["report"] == "commune":
        sheet_evolution(wb, rows, ci, f" — {scope}")
        sheet_par_site(wb, rows)
    elif p["report"] == "avant_apres":
        if not ci or not ci.get("travaux_debut"):
            raise SystemExit("Dates de travaux inconnues pour cette commune.")
        alloc_site = alloc_windows(invoices, periods, charges, d(ci["travaux_debut"]), d(ci["travaux_fin"]),
                                   lambda inv: (inv.get("sites") or {}).get("nom", "—"))
        sheet_avant_apres(wb, rows, alloc_site, ci, scope)
    else:  # synthese
        # Bande grise = fenêtre de travaux GLOBALE (min lancement → max achèvement des communes présentes).
        present = {i.get("commune_id") for i in invoices}
        debuts = [d(c["travaux_debut"]) for cid, c in communes_info.items()
                  if cid in present and c.get("travaux_debut")]
        fins = [d(c["travaux_fin"]) for cid, c in communes_info.items()
                if cid in present and c.get("travaux_fin")]
        global_ci = {"nom": "toutes communes", "travaux_debut": min(debuts).isoformat(),
                     "travaux_fin": max(fins).isoformat(), "travaux_estimes": False} if debuts and fins else None
        sheet_evolution(wb, rows, global_ci, "")
        sheet_synthese_avant_apres(wb, invoices, periods, charges, communes_info)

    tcd = wb.create_sheet("TCD")
    tcd["A1"] = "Tableau croisé dynamique (€ et kWh) — actualisé automatiquement à l'ouverture dans Excel."
    tcd["A1"].font = SUB_FONT
    sheet_datalogger(wb, p, scope, invoices, rows)
    ws_data = sheet_donnees(wb, rows)  # dernière position + masquée
    wb.save(out_path)

    try:
        inject_pivots(out_path, ws_data.max_row, [
            {"sheet_index": wb.sheetnames.index("TCD") + 1, "name": "TCD_Ability",
             "rows": [COL["Commune"], COL["Site"]], "cols": [COL["Période"]],
             "data": [(COL["Montant €"], "Somme de Montant €"), (COL["kWh"], "Somme de kWh")], "ref": "A3"},
        ])
    except Exception as exc:  # repli : classeur valide sans TCD
        sys.stderr.write(f"[pivot] injection ignorée : {exc}\n")

def main():
    if len(sys.argv) < 3:
        raise SystemExit("usage: generate_report.py '<params-json>' <out.xlsx>")
    build(json.loads(sys.argv[1]), sys.argv[2])
    print("OK")

if __name__ == "__main__":
    main()
